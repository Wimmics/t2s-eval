import json

from openpyxl import load_workbook


def extract_metrics():
    """Extract data from Excel and save as JSON"""
    # Load Excel file
    wb = load_workbook("Text2SPARQL Metrics.xlsx", data_only=True)
    ws = wb["Metrics"]

    data = []

    for row in ws.iter_rows(min_row=2, max_row=63, min_col=1, max_col=7):
        id = int(row[0].value)

        name = row[1].value
        definition = row[2].value
        used_in_ref = []
        if row[3].value is not None:
            for ref in str(row[3].value).split(","):
                ref = ref.strip()
                if ref != "":
                    used_in_ref.append(int(float(ref)))

        m_type = row[4].value
        reference = (
            int(row[5].value)
            if row[5].value is not None and row[5].value != "Link"
            else None
        )

        data.append(
            {
                "id": id,
                "name": name,
                "definition": definition,
                "used_in_ref": used_in_ref,
                "type": m_type,
                "reference": reference,
            }
        )

    with open("metrics.json", "w") as f:
        json.dump(data, f, indent=2)


def read_json():
    with open("metrics.json") as f:
        data = json.load(f)

    data.sort(key=lambda x: x["used_in_ref"].__len__(), reverse=True)
    for item in data:
        print(f"{item['name']} => {len(item['used_in_ref'])}")


def check_accepted_paper_used_metrics():
    with open("papers.json") as f:
        papers = json.load(f)

    with open("metrics.json") as f:
        metrics = json.load(f)

    for paper in papers:
        ref_id = paper["id"]
        used = False
        used_metrics = []
        for metric in metrics:
            if ref_id in metric["used_in_ref"]:
                used = True
                used_metrics.append(metric["name"])

        if not used:
            print(f"❌ Reference {ref_id} is not used in any metric.")
        # else:
        #     print(f"✅ Reference {ref_id} is used in {len(used_metrics)} metric(s)")


def check_metrics_ref_are_accepted():
    with open("papers.json") as f:
        papers = json.load(f)

    with open("metrics.json") as f:
        metrics = json.load(f)

    for metric in metrics:
        for ref_id in metric["used_in_ref"]:
            if not any(paper["id"] == ref_id for paper in papers):
                print(
                    f"❌ Reference {ref_id} used in metric {metric['name']} is not accepted."
                )
        # else:
        #     print(f"✅ Reference {ref_id} is used in {len(used_metrics)} metric(s)")
