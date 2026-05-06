import json

import bibtexparser
from openpyxl import load_workbook


def construct_paper_json():
    """Extract data from Excel and save as JSON."""
    # Load Excel file
    wb = load_workbook("Text2SPARQL Metrics.xlsx", data_only=True)
    ws = wb.active

    data = []

    for row in ws.iter_rows(min_row=2, max_row=142, min_col=1, max_col=14):
        accepted = row[6].value
        if accepted != "Yes":
            continue

        ref = int(row[0].value)

        title = row[1].value

        paper = row[2].hyperlink.target if row[2].hyperlink else None

        year = int(row[11].value)

        data.append({"title": title, "id": ref, "link": paper, "year": year})

    with open("papers.json", "w") as f:
        json.dump(data, f, indent=2)


def read_json():
    with open("papers.json") as f:
        data = json.load(f)

    data.sort(key=lambda x: x["title"])
    for item in data:
        print(item["title"])


def read_bibtex():
    with open("papers.bib") as f:
        bib_database = bibtexparser.load(f)

    for entry in bib_database.entries:
        print(entry["title"].strip())


def check_bibtex_json():
    with open("papers.bib") as f:
        bib_database = bibtexparser.load(f)

    with open("papers.json") as f:
        data = json.load(f)

    count = 0

    for item in data:
        title = item["title"]
        found = False
        for entry in bib_database.entries:
            if entry["title"].lower().strip() == title.lower().strip():
                found = True
                break
        if not found:
            print(f"❌ Not found in BibTeX: {item['id']} {title}")
            count += 1

    print(f"Total not found: {count}")


def check_bibtex_metrics_json():
    with open("metrics.bib") as f:
        bib_database = bibtexparser.load(f)

    for entry in bib_database.entries:
        print(f"\\cite{{{entry['ID']}}}")
