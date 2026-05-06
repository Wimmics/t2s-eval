"""Generate the static Text2SPARQL metrics inventory HTML page from the XLSX dataset."""

import json

import pandas as pd
from openpyxl import load_workbook

file_path = "datasets/Text2SPARQL_Metrics_Inventory.xlsx"
xls = pd.ExcelFile(file_path)
wb = load_workbook(file_path, data_only=True)

# Prepare data for HTML
html_data = {}
for sheet in xls.sheet_names:
    df = pd.read_excel(file_path, sheet_name=sheet)
    ws = wb[sheet]

    # If a cell has a hyperlink, use its target URL instead of display text.
    header_positions = {
        str(ws.cell(1, idx + 1).value): idx + 1 for idx in range(ws.max_column)
    }
    for col_name in df.columns:
        col_idx = header_positions.get(str(col_name))
        if not col_idx:
            continue
        for row_idx in range(len(df)):
            cell = ws.cell(row=row_idx + 2, column=col_idx)
            if cell.hyperlink and cell.hyperlink.target:
                df.at[row_idx, col_name] = cell.hyperlink.target

    # Convert dataframe to list of dicts
    html_data[sheet] = df.to_dict(orient="records")

# Define the HTML template
html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Text2SPARQL Metrics Inventory</title>
    <!-- Bootstrap CSS -->
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <!-- DataTables CSS -->
    <link href="https://cdn.datatables.net/1.13.6/css/dataTables.bootstrap5.min.css" rel="stylesheet">
    <style>
        body { background-color: #f8f9fa; padding: 20px; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
        .container-fluid { background: white; padding: 30px; border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
        h1 { color: #0d6efd; margin-bottom: 25px; font-weight: 700; }
        .nav-tabs { margin-bottom: 20px; border-bottom: 2px solid #dee2e6; }
        .nav-link { font-weight: 600; color: #495057; }
        .nav-link.active { color: #0d6efd !important; border-bottom: 3px solid #0d6efd !important; }
        table.dataTable thead { background-color: #f1f3f5; }
        .badge-ref { background-color: #e9ecef; color: #495057; font-weight: bold; }
        .link-cell { color: #0d6efd; text-decoration: none; font-weight: 500; }
        .link-cell:hover { text-decoration: underline; }
        .filter-row { background: #f8fafc; border: 1px solid #e6edf5; border-radius: 10px; padding: 12px; margin-bottom: 14px; }
        .chip-ref { cursor: pointer; }
        .chip-ref:hover { filter: brightness(0.95); }
        .table-note { font-size: 0.9rem; color: #5c6773; }
        tr.row-focus { outline: 2px solid #7bb4ff; outline-offset: -2px; }
    </style>
</head>
<body>

<div class="container-fluid">
    <h1>Text2SPARQL Metrics Inventory</h1>
    
    <ul class="nav nav-tabs" id="myTab" role="tablist">
        <li class="nav-item" role="presentation">
            <button class="nav-link active" id="papers-tab" data-bs-toggle="tab" data-bs-target="#papers" type="button" role="tab">Papers</button>
        </li>
        <li class="nav-item" role="presentation">
            <button class="nav-link" id="metrics-tab" data-bs-toggle="tab" data-bs-target="#metrics" type="button" role="tab">Metrics</button>
        </li>
    </ul>

    <div class="tab-content" id="myTabContent">
        <!-- Papers Tab -->
        <div class="tab-pane fade show active" id="papers" role="tabpanel">
            <div class="filter-row">
                <div class="row g-2 align-items-end">
                    <div class="col-md-3">
                        <label class="form-label mb-1" for="paperYearFilter">Year</label>
                        <select id="paperYearFilter" class="form-select form-select-sm">
                            <option value="">All years</option>
                        </select>
                    </div>
                    <div class="col-md-5">
                        <label class="form-label mb-1" for="paperVenueFilter">Venue</label>
                        <select id="paperVenueFilter" class="form-select form-select-sm">
                            <option value="">All venues</option>
                        </select>
                    </div>
                    <div class="col-md-4 text-md-end">
                        <button id="clearPaperFilters" class="btn btn-sm btn-outline-secondary">Clear Paper Filters</button>
                    </div>
                </div>
                <div id="refFilterHint" class="table-note mt-2">Tip: click references in Metrics table to jump here and focus matching papers.</div>
            </div>
            <div class="table-responsive">
                <table id="table-papers" class="table table-striped table-hover w-100">
                    <thead>
                        <tr>
                            <th>Ref</th>
                            <th>Title</th>
                            <th>Paper</th>
                            <th>Repo</th>
                            <th>Venue</th>
                            <th>Year</th>
                            <th>Peer Reviewed</th>
                        </tr>
                    </thead>
                    <tbody></tbody>
                </table>
            </div>
        </div>

        <!-- Metrics Tab -->
        <div class="tab-pane fade" id="metrics" role="tabpanel">
            <div class="filter-row">
                <div class="row g-2 align-items-end">
                    <div class="col-md-4">
                        <label class="form-label mb-1" for="metricTypeFilter">Metric Type</label>
                        <select id="metricTypeFilter" class="form-select form-select-sm">
                            <option value="">All types</option>
                        </select>
                    </div>
                    <div class="col-md-4">
                        <label class="form-label mb-1" for="metricRefSearch">Used In Ref contains</label>
                        <input id="metricRefSearch" class="form-control form-control-sm" placeholder="e.g. 38">
                    </div>
                    <div class="col-md-4 text-md-end">
                        <button id="clearMetricFilters" class="btn btn-sm btn-outline-secondary">Clear Metric Filters</button>
                    </div>
                </div>
            </div>
            <div class="table-responsive">
                <table id="table-metrics" class="table table-striped table-hover w-100">
                    <thead>
                        <tr>
                            <th>Ref</th>
                            <th>Name</th>
                            <th>Definition</th>
                            <th>Type</th>
                            <th>Used In Ref</th>
                        </tr>
                    </thead>
                    <tbody></tbody>
                </table>
            </div>
        </div>
    </div>
</div>

<!-- Scripts -->
<script src="https://code.jquery.com/jquery-3.7.0.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
<script src="https://cdn.datatables.net/1.13.6/js/dataTables.bootstrap5.min.js"></script>

<script>
    const data = JSON_DATA_HERE;

    function toArrayRefs(input) {
        if (input === undefined || input === null || input === '' || Number.isNaN(input)) {
            return [];
        }
        return String(input)
            .split(',')
            .map(x => x.trim())
            .filter(Boolean);
    }

    function safeText(value) {
        if (value === undefined || value === null || Number.isNaN(value)) {
            return '';
        }
        return String(value);
    }

    $(document).ready(function() {
        let paperRefFilterSet = null;

        $.fn.dataTable.ext.search.push(function(settings, _data, dataIndex) {
            if (settings.nTable.id !== 'table-papers') {
                return true;
            }
            const tableApi = new $.fn.dataTable.Api(settings);
            const row = tableApi.row(dataIndex).data();
            if (!row) {
                return true;
            }
            const yearFilter = $('#paperYearFilter').val();
            const venueFilter = $('#paperVenueFilter').val();

            const yearOk = !yearFilter || String(row.Year) === String(yearFilter);
            const venueOk = !venueFilter || String(row.Venue) === String(venueFilter);
            const refOk = !paperRefFilterSet || paperRefFilterSet.has(String(row.Ref));
            return yearOk && venueOk && refOk;
        });

        $.fn.dataTable.ext.search.push(function(settings, data) {
            if (settings.nTable.id !== 'table-metrics') {
                return true;
            }
            const typeFilter = $('#metricTypeFilter').val();
            const refSearch = $('#metricRefSearch').val().trim();
            const typeCell = String(data[3] || '');
            const usedInCell = String(data[4] || '');
            const typeOk = !typeFilter || typeCell === typeFilter;
            const refOk = !refSearch || usedInCell.includes(refSearch);
            return typeOk && refOk;
        });

        // Initialize Papers Table
        const papersTable = $('#table-papers').DataTable({
            data: data.Papers,
            columns: [
                { data: 'Ref', render: d => `<span class="badge badge-ref">${d}</span>` },
                { data: 'Title', className: 'fw-bold' },
                { data: 'Paper', render: d => (d === 'Link' || d === 'LInk') ? '<span class="text-muted">Link</span>' : `<a href="${d}" target="_blank" class="link-cell">View</a>` },
                { data: 'Repo', render: d => (d === 'Link' || !safeText(d)) ? '<span class="text-muted">Link</span>' : `<a href="${d}" target="_blank" class="link-cell">Repo</a>` },
                { data: 'Venue' },
                { data: 'Year' },
                { data: 'Peer Reviewed Type' }
            ],
            pageLength: 10,
            order: [[0, 'asc']]
        });

        // Initialize Metrics Table
        const metricsTable = $('#table-metrics').DataTable({
            data: data.Metrics,
            columns: [
                { data: 'Ref', render: d => `<span class="badge badge-ref">${d}</span>` },
                { data: 'Name', className: 'fw-bold text-primary' },
                { data: 'Definition', render: d => safeText(d) },
                { data: 'Type', render: d => safeText(d) },
                { data: 'Used In Ref', render: d => {
                    if (!d) return '';
                    return toArrayRefs(d)
                        .map(ref => `<span class="badge bg-light text-dark border me-1 chip-ref" data-ref="${ref}">${ref}</span>`)
                        .join('');
                }}
            ],
            pageLength: 10,
            order: [[0, 'asc']]
        });

        const years = [...new Set(data.Papers.map(p => p.Year).filter(v => v !== null && v !== undefined))].sort((a, b) => a - b);
        years.forEach(y => $('#paperYearFilter').append(`<option value="${y}">${y}</option>`));

        const venues = [...new Set(data.Papers.map(p => p.Venue).filter(Boolean))].sort((a, b) => String(a).localeCompare(String(b)));
        venues.forEach(v => $('#paperVenueFilter').append(`<option value="${v}">${v}</option>`));

        const metricTypes = [...new Set(data.Metrics.map(m => m.Type).filter(v => v && !Number.isNaN(v)))].sort((a, b) => String(a).localeCompare(String(b)));
        metricTypes.forEach(t => $('#metricTypeFilter').append(`<option value="${t}">${t}</option>`));

        $('#paperYearFilter, #paperVenueFilter').on('change', function() {
            papersTable.draw();
        });

        $('#metricTypeFilter').on('change', function() {
            metricsTable.draw();
        });

        $('#metricRefSearch').on('input', function() {
            metricsTable.draw();
        });

        $('#clearPaperFilters').on('click', function() {
            paperRefFilterSet = null;
            $('#paperYearFilter').val('');
            $('#paperVenueFilter').val('');
            $('#refFilterHint').text('Tip: click references in Metrics table to jump here and focus matching papers.');
            $('#table-papers tbody tr').removeClass('row-focus');
            papersTable.draw();
        });

        $('#clearMetricFilters').on('click', function() {
            $('#metricTypeFilter').val('');
            $('#metricRefSearch').val('');
            metricsTable.search('').draw();
        });

        $('#table-metrics tbody').on('click', '.chip-ref', function() {
            const ref = String($(this).data('ref'));
            paperRefFilterSet = new Set([ref]);
            $('#refFilterHint').text(`Paper filter active from Metrics: Ref ${ref}.`);
            papersTable.draw();

            const papersTab = new bootstrap.Tab(document.querySelector('#papers-tab'));
            papersTab.show();

            setTimeout(function() {
                $('#table-papers tbody tr').removeClass('row-focus');
                $('#table-papers tbody tr').each(function() {
                    const rowData = papersTable.row(this).data();
                    if (rowData && String(rowData.Ref) === ref) {
                        $(this).addClass('row-focus');
                        this.scrollIntoView({ behavior: 'smooth', block: 'center' });
                    }
                });
            }, 120);
        });

        $('#table-metrics tbody').on('click', 'tr', function(evt) {
            if ($(evt.target).closest('.chip-ref').length) {
                return;
            }
            const rowData = metricsTable.row(this).data();
            if (!rowData) {
                return;
            }
            const refs = toArrayRefs(rowData['Used In Ref']);
            if (!refs.length) {
                return;
            }
            paperRefFilterSet = new Set(refs);
            $('#refFilterHint').text(`Paper filter active from metric "${rowData.Name}": refs ${refs.join(', ')}.`);
            papersTable.draw();

            const papersTab = new bootstrap.Tab(document.querySelector('#papers-tab'));
            papersTab.show();
        });
    });
</script>

</body>
</html>
"""

# Inject data
full_html = html_template.replace("JSON_DATA_HERE", json.dumps(html_data))

with open("docs/text2sparql-metrics-inventory/index.html", "w", encoding="utf-8") as f:
    f.write(full_html)

print("HTML file generated successfully.")
